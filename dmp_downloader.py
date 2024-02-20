import requests
from openpyxl import load_workbook, Workbook
from datetime import date
import csv
from typing import TextIO

def download_dmp_data(month: str, year: int) -> Workbook:
    """Function to return latest DMP data file for processing"""

    url: str = f"https://www.bankofengland.co.uk/-/media/boe/files/decision-maker-panel-survey/{year}/monthly-dmp-data-for-website-{month}-{year}.xlsx"
    response: requests.Response = requests.get(url)

    if response.status_code == 200:
        output: TextIO = open(f"dmp_{month}_{year}.xlsx", "wb")
        output.write(response.content)
        output.close()

    wb: Workbook = load_workbook(f"dmp_{month}_{year}.xlsx")

    return wb

def months_diff(date_1: date, date_2: date) -> int:
    return (date_1.month - date_2.month) + 12 * (date_1.year - date_2.year)

def prepare_xlsx_file(
        wb: Workbook,
        ws_names: list[str],
        month: str,
        year: int
    ) -> dict[str, int]:
    """Get column configuration metadata and prepare .xlsx file for reading as .csv"""

    starting_row_dict: dict[str, int] = {ws_name: None for ws_name in ws_names}

    for ws_name in ws_names:
        ws = wb[ws_name]
        row_number: int = 1
        
        while ws["A"][row_number].number_format != "mmm-yy":
            row_number += 1

        ws.cell(row=row_number, column=1).value = "time"

        if ws_name == "Employment growth":
            ws.cell(row=row_number, column=2).value = "Mean Realised Employment Growth"
            ws.cell(row=row_number, column=13).value = "Mean Expected Employment Growth"
        starting_row_dict[ws_name] = row_number
    
    wb.save(f"dmp_{month}_{year}.xlsx")

    return starting_row_dict

def save_csv_file(
        wb: Workbook, 
        ws_names: list[str], 
        month: str,
        year: int,
        starting_row_dict: dict[str, int], 
        column_lengths_dict: dict[str, int]
    ) -> list[str]:
    """Write necessary .xlsx data into separate .csv files to be consumed by main Marko Chain functions"""

    output_file_paths: list[str] = []

    for ws_name in ws_names:
        ws = wb[ws_name]
        formatted_name: str = ws_name.replace(" ", "_").lower()
        output_file_path: str = f"{formatted_name}_{month}_{year}.csv"
        with open(output_file_path, "w", newline="") as f:
            c = csv.writer(f)
            row_count: int = 0
            write_count: int = 0
            for row in ws.rows:
                if row_count >= starting_row_dict[ws_name] - 1:
                    c.writerow([cell.value for cell in row if cell.column_letter in ["A", "B", "M"]] if ws_name == "Employment growth" else [cell.value for cell in row])
                    write_count += 1
                row_count += 1
                if write_count > column_lengths_dict[ws_name] + 1:
                    break
        
        output_file_paths.append(output_file_path)
    
    return output_file_paths

def download_and_process_dmp_data(
        month: str,
        year: int,
        ws_names: list[str],
        column_lengths_dict: dict[str, int]
    ) -> list[str]:
    """Main function to ingest the DMP data by downloading and cleaning the necessary columns and saving out the .csv files"""

    wb: Workbook = download_dmp_data(month, year)
    starting_row_dict: dict[str, int] = prepare_xlsx_file(wb, ws_names, month, year)
    output_file_paths: list[str] = save_csv_file(wb, ws_names, month, year, starting_row_dict, column_lengths_dict)

    return output_file_paths