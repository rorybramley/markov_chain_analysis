from openpyxl import Workbook
from typing import List
from pandas import DataFrame

def generate_flat_result_dict(
        means_dict: dict[str, float], 
        std_devs_dict,
        dfs_dict: dict[str, List[DataFrame]],
    ) -> dict[str, dict[str, float]]:

    flat_results_dict: dict[str, dict[str, float]] = {}

    for key, value in dfs_dict.items():
        flat_results_dict[f"{key}_up_one_iteration"] = value[0]["up"][0]
        flat_results_dict[f"{key}_down_one_iteration"] = value[0]["down"][0]
        flat_results_dict[f"{key}_up_nth_iteration"] = value[-1]["up"][0]
        flat_results_dict[f"{key}_down_nth_iteration"] = value[-1]["down"][0]
        flat_results_dict[f"{key}_mean"] = means_dict[key]
        flat_results_dict[f"{key}_std_dev"] = std_devs_dict[key]
    
    return flat_results_dict

def write_results_to_xlsx(
        input_file_paths: List[str], 
        results: List[tuple[List[float], List[float], dict[str, List[DataFrame]]]],
        output_path: str):
    """Take in results of Markov Chains and add each set to its own sheet in a results .xlsx file"""

    sheet_names: List[str] = [name.split(".")[0] for name in input_file_paths]
    output_dict: dict[str, dict[str, float]] = {sheet_name: generate_flat_result_dict(*result) for sheet_name, result in zip(sheet_names, results)}

    wb: Workbook = Workbook()

    for index, sheet_name in enumerate(sheet_names):
        wb.create_sheet(title=sheet_name, index=index)
    
    wb.remove(wb["Sheet"])

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        data = output_dict[sheet_name]
        header = list(data.keys())
        ws.append(header)

        cell = ws.cell(row=2, column=1)
        for value in data.values():
            cell.value = value 
            cell = ws.cell(row=2, column=cell.column+1)
        
    wb.save(output_path)

    print(f"File saved to {output_path} succesfully!")